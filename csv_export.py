#!/usr/bin/env python3

import argparse
import csv
import json
import os
import re
import sys
import shutil
from openpyxl import load_workbook
from numpy import *

# offset by 3
met_index = {
    "PSNR": 0,
    "PSNRHVS": 1,
    "SSIM": 2,
    "FASTSSIM": 3,
    "CIEDE2000": 4,
    "PSNR Cb": 5,
    "PSNR Cr": 6,
    "APSNR": 7,
    "APSNR Cb": 8,
    "APSNR Cr": 9,
    "MSSSIM": 10,
    "Encoding Time": 11,
    "VMAF_old": 12,
    "Decoding Time": 13,
    "PSNR Y (libvmaf)": 14,
    "PSNR Cb (libvmaf)": 15,
    "PSNR Cr (libvmaf)": 16,
    "CIEDE2000 (libvmaf)": 17,
    "SSIM (libvmaf)": 18,
    "MS-SSIM (libvmaf)": 19,
    "PSNR-HVS Y (libvmaf)": 20,
    "PSNR-HVS Cb (libvmaf)": 21,
    "PSNR-HVS Cr (libvmaf)": 22,
    "PSNR-HVS (libvmaf)": 23,
    "VMAF": 24,
    "VMAF-NEG": 25,
    "APSNR Y (libvmaf)": 26,
    "APSNR Cb (libvmaf)": 27,
    "APSNR Cr (libvmaf)": 28,
    "CAMBI (libvmaf)": 29,
}

# row_id for different sets inside template.
start_rows = {
    'A1': 2,
    'A2': 50,
    'A3': 182,
    'A4': 230,
    'A5': 266,
    'B1': 290,
    'B2': 350,
    'G1': 416,
    'G2': 440,
    'E': 482
}

# CTC Configs
# LD : ctc_sets_mandatory
# RA: ctc_sets_mandatory  + ctc_sets_optional
# AI: ctc_sets_mandatory_ai + ctc_sets_optional
# AS: A1 with Downsampling
ctc_sets_mandatory = [
    "aomctc-a1-4k",
    "aomctc-a2-2k",
    "aomctc-a3-720p",
    "aomctc-a4-360p",
    "aomctc-a5-270p",
    "aomctc-b1-syn",
    "aomctc-b2-syn"]
ctc_sets_mandatory_ai = ctc_sets_mandatory + \
    ["aomctc-f1-hires", "aomctc-f2-midres"]
ctc_sets_optional = ["aomctc-g1-hdr-4k",
                     "aomctc-g2-hdr-2k", "aomctc-e-nonpristine"]

run_cfgs = ['RA', 'LD', 'AI', 'AS']

quality_presets = {
    "daala": [7, 11, 16, 25, 37],
    "x264": list(range(1, 52, 5)),
    "x265": list(range(5, 52, 5)),
    "x265-rt": list(range(5, 52, 5)),
    "xvc": [20, 25, 30, 35, 40],
    "vp8": list(range(12, 64, 5)),
    "vp9": [20, 32, 43, 55, 63],
    "vp9-rt": [20, 32, 43, 55, 63],
    "vp10": [8, 20, 32, 43, 55, 63],
    "vp10-rt": [8, 20, 32, 43, 55, 63],
    "av1": [20, 32, 43, 55, 63],
    "av1-rt": [20, 32, 43, 55, 63],
    "av2-ai": [85, 110, 135, 160, 185, 210],
    "av2-ra": [110, 135, 160, 185, 210, 235],
    "av2-ra-st": [110, 135, 160, 185, 210, 235],
    "av2-ld": [110, 135, 160, 185, 210, 235],
    "av2-as": [110, 135, 160, 185, 210, 235],
    "av2-f": [60, 85, 110, 135, 160, 185],
    "thor": list(range(7, 43, 3)),
    "thor-rt": list(range(7, 43, 3)),
    "rav1e": [20 * 4, 32 * 4, 43 * 4, 55 * 4, 63 * 4],
    "svt-av1": [20, 32, 43, 55, 63],
}

row_header = [
    "TestCfg",
    "EncodeMethod",
    "CodecName",
    "EncodePreset",
    "Class",
    "Name",
    "OrigRes",
    "FPS",
    "BitDepth",
    "CodedRes",
    "QP",
    "Bitrate(kbps)",
    "PSNR_Y",
    "PSNR_U",
    "PSNR_V",
    "SSIM_Y(dB)",
    "MS-SSIM_Y(dB)",
    "VMAF_Y",
    "VMAF_Y-NEG",
    "PSNR-HVS",
    "CIEDE2000",
    "APSNR_Y",
    "APSNR_U",
    "APSNR_V",
    "CAMBI",
    "EncT[s]",
    "DecT[s]",
    "EncInstr",
    "DecInstr",
    "EncCycles",
    "DecCycles",
    "EncMD5"
]


class Logger(object):
    def __init__(self, run_path, args):
        self.this_args = args
        if not args.ctc_export:
            self.terminal = sys.stdout
        self.log = open(run_path + "/csv_export.csv", "w")

    def write(self, message):
        if not self.this_args.ctc_export:
            self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        if not self.this_args.ctc_export:
            self.terminal.flush()
        self.log.flush()


def return_start_rows(set_name):
    try:
        if 'aomctc' in set_name:
            normalized_set = set_name.split('-')[1].upper()
            if normalized_set in start_rows.keys():
                return start_rows[normalized_set], normalized_set
    except BaseException:
        print("Not a CTC set to Normalize, check the runs")
        sys.exit(1)


def return_ctc_set_list(run_info):
    set_name = run_info['ctcSets']
    config = run_info['codec']
    if 'aomctc-all' in set_name:
        if config == 'av2-ai':
            run_set_list = ctc_sets_mandatory_ai + ctc_sets_optional
        elif config == 'av2-ra-st' or config == 'av2-ra':
            run_set_list = ctc_sets_mandatory + ctc_sets_optional
        elif config == 'av2-ld':
            run_set_list = ctc_sets_mandatory
    elif 'aomctc-mandatory' in set_name:
        if config == 'av2-ra-st' or config == 'av2-ra' or config == 'av2-ld':
            run_set_list = ctc_sets_mandatory
        elif config == 'av2-ai':
            run_set_list = ctc_sets_mandatory_ai
    else:
        run_set_list = run_info['ctcSets']
    return run_set_list


def write_set_data(run_path, writer, current_video_set):
    info_data = json.load(open(run_path + "/info.json"))
    videos_dir = os.path.join(
        os.getenv("MEDIAS_SRC_DIR", "/mnt/runs/sets"), current_video_set
    )  # for getting framerate
    sets = json.load(
        open(os.path.join(os.getenv("CONFIG_DIR", "rd_tool"), "sets.json")))
    videos = sets[current_video_set]["sources"]
    # sort name ascending, resolution descending
    if current_video_set != "av2-a1-4k-as":
        videos.sort(key=lambda s: s.lower())
    else:
        videos.sort(
            key=lambda x: x.split("_")[0]
            + "%08d" % (100000 - int(x.split("_")[1].split("x")[0]))
        )
    if 'av2' in info_data['codec']:
        normalized_cfg = info_data['codec'].split('-')[1].upper()
    else:
        normalized_cfg = 'RA'
    # Get the Quality values, if user defined, use that, else do defaults
    if 'qualities' in list(info_data.keys()):
        qp_list = info_data['qualities'].split()
    else:
        qp_list = quality_presets[info_data['codec']]
    try:
        for video in videos:
            v = open(os.path.join(videos_dir, video), "rb")
            line = v.readline().decode("utf-8")
            fps_n, fps_d = re.search(r"F([0-9]*)\:([0-9]*)", line).group(1, 2)
            width = re.search(r"W([0-9]*)", line).group(1)
            height = re.search(r"H([0-9]*)", line).group(1)
            if 'aomctc' in current_video_set:
                normalized_set = current_video_set.split('-')[1].upper()
            a = loadtxt(
                os.path.join(
                    run_path,
                    current_video_set,
                    video +
                    "-daala.out"))
            # This way, even partial information from the *daala.out can be
            # rendered by having key-value where key is QP.
            encoded_qp_list = {}
            for row in a:
                encoded_qp_list[int(row[0])] = row
            for this_qp in qp_list:
                # Check if the QPs is present in the currently stored daala.out
                if this_qp in encoded_qp_list.keys():
                    row = encoded_qp_list[this_qp]
                    frames = int(row[1]) / int(width) / int(height)
                    if info_data["codec"] == "av2-as":
                        writer.writerow(
                            [
                                "AS",  # TestCfg
                                "aom",  # EncodeMethod
                                info_data["run_id"],  # CodecName
                                "",  # EncodePreset
                                normalized_set,  # Class
                                video,  # name
                                "3840x2160",  # OrigRes
                                "",  # FPS
                                10,  # BitDepth
                                str(width) + "x" + str(height),  # CodedRes
                                row[0],  # qp
                                int(row[2])
                                * 8.0
                                * float(fps_n)
                                / float(fps_d)
                                / frames
                                / 1000.0,  # bitrate
                                row[met_index["PSNR Y (libvmaf)"] + 3],
                                row[met_index["PSNR Cb (libvmaf)"] + 3],
                                row[met_index["PSNR Cr (libvmaf)"] + 3],
                                row[met_index["SSIM (libvmaf)"] + 3],
                                row[met_index["MS-SSIM (libvmaf)"] + 3],
                                row[met_index["VMAF"] + 3],
                                row[met_index["VMAF-NEG"] + 3],
                                row[met_index["PSNR-HVS Y (libvmaf)"] + 3],
                                row[met_index["CIEDE2000 (libvmaf)"] + 3],
                                row[met_index["APSNR Y (libvmaf)"] + 3],
                                row[met_index["APSNR Cb (libvmaf)"] + 3],
                                row[met_index["APSNR Cr (libvmaf)"] + 3],
                                row[met_index["Encoding Time"] + 3],
                                row[met_index["Decoding Time"] + 3],
                            ]
                        )
                    else:
                        writer.writerow(
                            [
                                normalized_cfg,  # TestCfg
                                "aom",  # EncodeMethod
                                info_data["run_id"],  # CodecName
                                0,  # EncodePreset #TODO: FIXME
                                normalized_set,  # Class
                                video,  # name
                                str(width) + "x" + str(height),  # OrigRes
                                str(float(fps_n) / float(fps_d)),  # FPS
                                10,  # BitDepth #TODO: FIXME
                                str(width) + "x" + str(height),  # CodedRes
                                int(row[0]),  # qp
                                int(row[2])
                                * 8.0
                                * float(fps_n)
                                / float(fps_d)
                                / frames
                                / 1000.0,  # bitrate
                                row[met_index["PSNR Y (libvmaf)"] + 3],
                                row[met_index["PSNR Cb (libvmaf)"] + 3],
                                row[met_index["PSNR Cr (libvmaf)"] + 3],
                                row[met_index["SSIM (libvmaf)"] + 3],
                                row[met_index["MS-SSIM (libvmaf)"] + 3],
                                row[met_index["VMAF"] + 3],
                                row[met_index["VMAF-NEG"] + 3],
                                row[met_index["PSNR-HVS Y (libvmaf)"] + 3],
                                row[met_index["CIEDE2000 (libvmaf)"] + 3],
                                row[met_index["APSNR Y (libvmaf)"] + 3],
                                row[met_index["APSNR Cb (libvmaf)"] + 3],
                                row[met_index["APSNR Cr (libvmaf)"] + 3],
                                row[met_index["CAMBI (libvmaf)"] + 3],
                                row[met_index["Encoding Time"] + 3],
                                row[met_index["Decoding Time"] + 3],
                                "",  # ENCInstr
                                "",  # DecInstr
                                "",  # EncCycles
                                "",  # DecCycles
                                "",  # EncMD5
                            ]
                        )
                # Case where the data is yet to be made
                else:
                    writer.writerow([
                        "RA",  # TestCfg # TODO: FIXME
                        "aom",  # EncodeMethod
                        info_data["run_id"],  # CodecName
                        0,  # EncodePreset #TODO: FIXME
                        normalized_set,  # Class
                        video,  # name
                        str(width) + "x" + str(height),  # OrigRes
                        str(float(fps_n) / float(fps_d)),  # FPS
                        10,  # BitDepth #TODO: FIXME
                        str(width) + "x" + str(height),  # CodedRes
                        this_qp  # qp
                    ])
    except BaseException:
        # This allows partial rendering of CSV + XLS Reports
        pass


def save_ctc_export(run_path, cmd_args):
    info_data = json.load(open(run_path + "/info.json"))
    task = info_data["task"]
    sets = json.load(
        open(os.path.join(os.getenv("CONFIG_DIR", "rd_tool"), "sets.json")))
    videos = sets[task]["sources"]
    # sort name ascending, resolution descending
    if task != "av2-a1-4k-as":
        videos.sort(key=lambda s: s.lower())
    else:
        videos.sort(
            key=lambda x: x.split("_")[0]
            + "%08d" % (100000 - int(x.split("_")[1].split("x")[0]))
        )
    if not cmd_args.ctc_export:
        sys.stdout = Logger(run_path, cmd_args)
        w = csv.writer(sys.stdout, dialect="excel")
        w.writerow(row_header)
        write_set_data(run_path, w, task)
    else:
        ctc_set_list = return_ctc_set_list(info_data)
        csv_writer_obj = open(run_path + "/csv_export.csv", 'w')
        w = csv.writer(csv_writer_obj, dialect="excel")
        w.writerow(row_header)
        # Abstract Writing per-set data
        for set_name in ctc_set_list:
            write_set_data(run_path, w, set_name)
        csv_writer_obj.close()


def write_xls_rows(run_path, current_video_set, this_sheet):
    run_file = open(run_path + '/csv_export.csv', 'r')
    start_id, normalized_set = return_start_rows(current_video_set)

    run_reader = csv.reader(run_file)
    next(run_reader)
    this_row = start_id
    for this_line in run_reader:
        if this_line[4] != normalized_set:
            continue
        else:
            this_col = 1
            for this_values in this_line:
                this_cell = this_sheet.cell(row=this_row, column=this_col)
                if this_col >= 12 and this_col <= 30 and this_values != "":
                    this_cell.value = float(this_values)
                else:
                    this_cell.value = this_values
                this_col += 1
        this_row += 1
    run_file.close()


def write_xls_file(run_a, run_b):
    xls_template = os.path.join(
        os.getenv("CONFIG_DIR", "rd_tool"), 'AOM_CWG_Regular_CTCv3_v7.2.xlsm')
    run_a_info = json.load(open(run_a + "/info.json"))
    run_b_info = json.load(open(run_b + "/info.json"))
    run_id_a = run_a_info["run_id"]
    run_id_b = run_b_info["run_id"]
    xls_file = run_a + '/../ctc_results/' + \
        "CTC_Regular_v0-%s-%s.xlsm" % (run_id_a, run_id_b)
    shutil.copyfile(xls_template, xls_file)
    wb = load_workbook(xls_file, read_only=False, keep_vba=True)
    this_codec = run_a_info['codec']
    if '-' in this_codec:
        if this_codec.split('-')[1].upper() in run_cfgs:
            this_cfg = this_codec.split('-')[1].upper()
    else:
        this_cfg = 'RA'
    anchor_sheet_name = 'Anchor-%s' % this_cfg
    anchor_sheet = wb[anchor_sheet_name]
    test_sheet_name = 'Test-%s' % this_cfg
    test_sheet = wb[test_sheet_name]
    current_video_set = run_a_info["task"]
    current_ctc_list_a = return_ctc_set_list(run_a_info)
    current_ctc_list_b = return_ctc_set_list(run_b_info)

    # Single Video Set Condition
    if len(current_ctc_list_a) == 0 and len(current_ctc_list_a) == 0:
        write_xls_rows(run_a, current_video_set, anchor_sheet)
        write_xls_rows(run_b, current_video_set, test_sheet)
        wb.save(xls_file)
    # Multi-Set Case
    else:
        for this_video_set in current_ctc_list_a:
            write_xls_rows(run_a, this_video_set, anchor_sheet)
        for this_video_set in current_ctc_list_b:
            write_xls_rows(run_b, this_video_set, test_sheet)
        wb.save(xls_file)


def main():
    parser = argparse.ArgumentParser(
        description="Generate CTC CSV version of .out files")
    parser.add_argument("run", nargs=1, help="Run folder (Anchor)")
    parser.add_argument("--ctc_export", action='store_true', help="XLS Export")
    parser.add_argument(
        "--run_b", help="Target Run Dir (Only used in CTC case)")
    args = parser.parse_args()

    if not args.ctc_export:
        save_ctc_export(args.run[0], args)
    else:
        if not args.run_b:
            print("ERROR: Missing Target, aborting")
            sys.exit(1)
        save_ctc_export(args.run[0], args)
        save_ctc_export(args.run_b, args)

        write_xls_file(args.run[0], args.run_b)


if __name__ == "__main__":
    main()
